"""Parse APQC PCF 7.4 Excel Combined sheet → framework.json + sources_mapping.json.

Reads the Combined sheet (1921 entries), builds a tree structure with
Blueprint v1.0 five-pillar fields, and outputs:
  - docs/oprocess-framework/framework.json (tree, English only in Phase 1)
  - docs/oprocess-framework/sources_mapping.json (PCF hierarchy_id → OPF id)
"""

from __future__ import annotations

import sys
import warnings
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

# Allow running from project root
sys.path.insert(0, str(Path(__file__).resolve().parent))

from shared.io import IdRegistry, compute_level, get_parent_id, write_json
from shared.text import normalize_text
from shared.types import LocalizedText, ProcessNode

# ── Constants ──────────────────────────────────────────────────────────

PCF_PATH = Path("docs/K014749_APQC Process Classification Framework (PCF)"
                " - Cross-Industry - Excel Version 7.4.xlsx")
OUTPUT_DIR = Path("docs/oprocess-framework")
FRAMEWORK_PATH = OUTPUT_DIR / "framework.json"
SOURCES_PATH = OUTPUT_DIR / "sources_mapping.json"

# PCF Categories 1-5 + 6 (Customer Service) = operating
OPERATING_CATEGORIES = {"1", "2", "3", "4", "5", "6"}

# Column names in the Combined sheet header row
COL_HIERARCHY_ID = "Hierarchy ID"
COL_NAME = "Name"
COL_DESCRIPTION = "Element Description"

# Tag keywords per L1 category number
CATEGORY_TAGS: dict[str, list[str]] = {
    "1": ["strategy"], "2": ["product", "service"],
    "3": ["marketing", "sales"], "4": ["supply_chain", "delivery"],
    "5": ["service_delivery"], "6": ["customer_service"],
    "7": ["hr", "human_capital"], "8": ["it_management"],
    "9": ["finance"], "10": ["asset_management"],
    "11": ["risk", "compliance"], "12": ["external_relations"],
    "13": ["business_capabilities"],
}


def _get_category_num(hierarchy_id: str) -> str:
    """Extract category number from hierarchy ID (e.g., '4.4.3' → '4')."""
    return hierarchy_id.split(".")[0]


def _get_domain(hierarchy_id: str) -> str:
    cat = _get_category_num(hierarchy_id)
    return "operating" if cat in OPERATING_CATEGORIES else "management_support"


def _build_tags(hierarchy_id: str) -> list[str]:
    cat = _get_category_num(hierarchy_id)
    tags = CATEGORY_TAGS.get(cat, []).copy()
    tags.append("pcf")
    return tags


def _read_combined_sheet(pcf_path: Path) -> list[dict[str, str]]:
    """Read PCF Combined sheet, returning list of row dicts."""
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        wb = load_workbook(str(pcf_path), read_only=True, data_only=True)

    try:
        ws = wb["Combined"]
        header: list[str] = []
        rows: list[dict[str, str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [normalize_text(str(c or "")) for c in row]
                continue
            row_dict = {}
            for j, cell in enumerate(row):
                if j < len(header):
                    row_dict[header[j]] = normalize_text(str(cell or ""))
            rows.append(row_dict)
        return rows
    finally:
        wb.close()


def _rows_to_nodes(
    rows: list[dict[str, str]], registry: IdRegistry
) -> list[ProcessNode]:
    """Convert row dicts to ProcessNode list, registering all IDs."""
    nodes: list[ProcessNode] = []
    for row in rows:
        hid = row[COL_HIERARCHY_ID]
        name_en = row[COL_NAME]
        desc_en = row.get(COL_DESCRIPTION, "")
        level = compute_level(hid)
        parent = get_parent_id(hid)

        registry.register(hid)

        node = ProcessNode(
            id=hid,
            level=level,
            parent_id=parent,
            domain=_get_domain(hid),
            source=[f"PCF:{hid}"],
            name=LocalizedText(en=name_en, zh=""),
            description=LocalizedText(en=desc_en, zh=""),
            tags=_build_tags(hid),
            ai_context=desc_en[:200] if desc_en else "",
        )
        nodes.append(node)
    return nodes


def _build_tree(flat_nodes: list[ProcessNode]) -> list[ProcessNode]:
    """Build tree from flat node list via dict lookup."""
    lookup: dict[str, ProcessNode] = {n.id: n for n in flat_nodes}
    roots: list[ProcessNode] = []
    for node in flat_nodes:
        if node.parent_id is None:
            roots.append(node)
        elif parent := lookup.get(node.parent_id):
            parent.children.append(node)
        else:
            raise ValueError(
                f"Orphan node {node.id}: parent {node.parent_id} not found"
            )
    return roots


def _build_sources_mapping(flat_nodes: list[ProcessNode]) -> dict[str, str]:
    """Build PCF hierarchy_id → OPF id mapping (identity for Phase 1)."""
    return {n.id: n.id for n in flat_nodes}


def main() -> None:
    print(f"Reading PCF Excel: {PCF_PATH}")
    rows = _read_combined_sheet(PCF_PATH)
    print(f"  Combined sheet: {len(rows)} rows")

    registry = IdRegistry()
    flat_nodes = _rows_to_nodes(rows, registry)
    print(f"  Registered IDs: {registry.count}")

    categories = _build_tree(flat_nodes)
    total = sum(c.count_nodes() for c in categories)
    print(f"  Tree: {len(categories)} categories, {total} total nodes")

    framework = {
        "version": "1.0.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_nodes": total,
        "categories": [c.to_dict() for c in categories],
    }
    write_json(framework, FRAMEWORK_PATH)
    print(f"  Written: {FRAMEWORK_PATH} ({FRAMEWORK_PATH.stat().st_size:,} bytes)")

    sources = _build_sources_mapping(flat_nodes)
    write_json(sources, SOURCES_PATH)
    print(f"  Written: {SOURCES_PATH} ({len(sources)} entries)")


if __name__ == "__main__":
    main()
