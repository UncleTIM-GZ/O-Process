"""Parse APQC PCF 7.4 Metrics sheet → kpis.json + update framework.json kpi_refs.

Reads the Metrics sheet (3910 entries), converts to KPI structure,
then links KPI refs back into framework.json nodes.
"""

from __future__ import annotations

import sys
import warnings
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from shared.io import read_json, write_json
from shared.text import normalize_text
from shared.types import KPIEntry, LocalizedText

# ── Constants ──────────────────────────────────────────────────────────

PCF_PATH = Path("docs/K014749_APQC Process Classification Framework (PCF)"
                " - Cross-Industry - Excel Version 7.4.xlsx")
OUTPUT_DIR = Path("docs/oprocess-framework")
KPIS_PATH = OUTPUT_DIR / "kpis.json"
FRAMEWORK_PATH = OUTPUT_DIR / "framework.json"

# Column indices in Metrics sheet
COL_HID = 1        # Hierarchy ID
COL_CATEGORY = 3   # Metric Category
COL_METRIC_ID = 4  # Metric ID
COL_NAME = 5       # Metric name
COL_FORMULA = 6    # Formula
COL_UNITS = 7      # Units


def _read_metrics_sheet(pcf_path: Path) -> list[KPIEntry]:
    """Read PCF Metrics sheet, returning list of KPIEntry."""
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        wb = load_workbook(str(pcf_path), read_only=True, data_only=True)

    try:
        ws = wb["Metrics"]
        kpis: list[KPIEntry] = []
        seq_counter: dict[str, int] = defaultdict(int)

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header

            hid = normalize_text(str(row[COL_HID] or ""))
            if not hid:
                continue

            seq_counter[hid] += 1
            seq = seq_counter[hid]

            name_en = normalize_text(str(row[COL_NAME] or ""))
            formula = normalize_text(str(row[COL_FORMULA] or "")) or None
            unit = normalize_text(str(row[COL_UNITS] or ""))
            category = normalize_text(str(row[COL_CATEGORY] or "")) or None

            # Handle null units
            if not unit or unit == "None":
                unit = "unknown"

            kpi = KPIEntry(
                id=f"kpi.{hid}.{seq:02d}",
                process_id=hid,
                name=LocalizedText(en=name_en, zh=""),
                unit=unit,
                formula=formula,
                category=category,
            )
            kpis.append(kpi)

        return kpis
    finally:
        wb.close()


def _link_kpi_refs(framework: dict, kpis: list[KPIEntry]) -> int:
    """Add kpi_refs to framework nodes. Returns count of linked refs."""
    # Build process_id → kpi_id mapping
    refs_by_process: dict[str, list[str]] = defaultdict(list)
    for kpi in kpis:
        refs_by_process[kpi.process_id].append(kpi.id)

    linked = 0

    def _update_node(node: dict) -> None:
        nonlocal linked
        node_id = node["id"]
        if node_id in refs_by_process:
            node["kpi_refs"] = refs_by_process[node_id]
            linked += len(refs_by_process[node_id])
        for child in node.get("children", []):
            _update_node(child)

    for cat in framework.get("categories", []):
        _update_node(cat)

    return linked


def main() -> None:
    print(f"Reading Metrics: {PCF_PATH}")
    kpis = _read_metrics_sheet(PCF_PATH)
    print(f"  Metrics: {len(kpis)} KPI entries")

    kpis_data = [kpi.to_dict() for kpi in kpis]
    write_json(kpis_data, KPIS_PATH)
    print(f"  Written: {KPIS_PATH} ({KPIS_PATH.stat().st_size:,} bytes)")

    # Link KPI refs into framework.json
    print(f"Linking KPI refs into {FRAMEWORK_PATH}")
    framework = read_json(FRAMEWORK_PATH)
    linked = _link_kpi_refs(framework, kpis)
    write_json(framework, FRAMEWORK_PATH)
    print(f"  Linked {linked} KPI refs across framework nodes")
    print(f"  Updated: {FRAMEWORK_PATH} ({FRAMEWORK_PATH.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
